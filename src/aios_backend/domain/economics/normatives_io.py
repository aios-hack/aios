from __future__ import annotations

import re
import zipfile
from pathlib import Path
from typing import Any

from aios_backend.core.contracts import EspCatalogEntry, NormativeSet

NORMATIVES_SHEET: str = "Нормативы"
ESP_SHEET: str = "ЭЦН"

RUB_PER_MILLION: float = 1_000_000.0

SCALAR_CODES: dict[str, str] = {
    "oilPriceRubT": "price_oil_rub_per_t",
    "deductionsRubT": "deductions_rub_per_t",
    "oilOpexRubT": "opex_oil_rub_per_t",
    "liquidOpexRubT": "opex_liquid_rub_per_t",
    "injectionOpexRubM3": "opex_injection_rub_per_m3",
    "fundAnnualRubWell": "opex_wellstock_rub_per_well_year",
    "pumpOperationCostM": "esp_swap_opex_rub",
    "stopStartCostM": "event_cost_rub",
    "conversionBaseCostM": "conversion_base_cost_rub",
    "waccRate": "wacc",
    "propertyTaxRate": "property_tax_rate",
    "profitTaxRate": "income_tax_rate",
}

MILLION_CODES: frozenset[str] = frozenset(
    {"pumpOperationCostM", "stopStartCostM", "conversionBaseCostM"}
)

PERCENT_CODES: frozenset[str] = frozenset(
    {"waccRate", "propertyTaxRate", "profitTaxRate"}
)

METHODOLOGY_LOCKED_RUB: dict[str, float] = {
    "event_cost_rub": 1_000_000.0,
    "conversion_base_cost_rub": 5_000_000.0,
}


class NormativesError(ValueError):
    pass


def _to_float(value: Any, code: str) -> float:
    if value is None or value == "":
        raise NormativesError(f"норматив {code}: пустое значение")
    if isinstance(value, bool):
        raise NormativesError(f"норматив {code}: логическое значение вместо числа")
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(" ", "").replace(" ", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        raise NormativesError(f"норматив {code}: не число, получено {value!r}") from None


def _scale(code: str, raw: float) -> float:
    if code in MILLION_CODES:
        return raw * RUB_PER_MILLION
    if code in PERCENT_CODES:
        return raw / 100.0
    return raw


def _read_sheets_via_openpyxl(path: Path) -> dict[str, list[list[Any]]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheets: dict[str, list[list[Any]]] = {}
        for name in (NORMATIVES_SHEET, ESP_SHEET):
            if name not in workbook.sheetnames:
                raise NormativesError(f"в файле нормативов нет листа «{name}»")
            sheets[name] = [
                list(row) for row in workbook[name].iter_rows(values_only=True)
            ]
        return sheets
    finally:
        workbook.close()


_CELL_RE = re.compile(rb'<c r="([A-Z]+)(\d+)"([^>]*)>(.*?)</c>|<c r="([A-Z]+)(\d+)"([^>]*)/>')
_VALUE_RE = re.compile(rb"<v>(.*?)</v>")
_INLINE_RE = re.compile(rb"<is>.*?</is>", re.S)
_TEXT_RE = re.compile(rb"<t[^>]*>(.*?)</t>", re.S)
_SHEET_RE = re.compile(rb"<sheet\s([^>]*)/?>")
_REL_RE = re.compile(rb"<Relationship\s([^>]*)/?>")
_ATTR_RE = re.compile(rb'([\w:]+)="([^"]*)"')
_SI_RE = re.compile(rb"<si>(.*?)</si>", re.S)
_ROW_RE = re.compile(rb"<row[^>]*>(.*?)</row>|<row[^>]*/>", re.S)


def _unescape(raw: bytes) -> str:
    text = raw.decode("utf-8")
    for entity, char in (
        ("&lt;", "<"),
        ("&gt;", ">"),
        ("&quot;", '"'),
        ("&apos;", "'"),
        ("&amp;", "&"),
    ):
        text = text.replace(entity, char)
    return text


def _resolve_member(target: str) -> str:
    normalized = target.replace("\\", "/")
    if normalized.startswith("/"):
        return normalized.lstrip("/")
    if normalized.startswith("xl/"):
        return normalized
    return f"xl/{normalized}"


def _attributes(raw: bytes) -> dict[str, str]:
    return {
        match.group(1).decode("utf-8"): _unescape(match.group(2))
        for match in _ATTR_RE.finditer(raw)
    }


def _column_index(letters: str) -> int:
    index = 0
    for char in letters:
        index = index * 26 + (ord(char) - ord("A") + 1)
    return index - 1


def _read_sheets_via_zip(path: Path) -> dict[str, list[list[Any]]]:
    with zipfile.ZipFile(path) as archive:
        workbook_xml = archive.read("xl/workbook.xml")
        rels_xml = archive.read("xl/_rels/workbook.xml.rels")
        targets: dict[str, str] = {}
        for match in _REL_RE.finditer(rels_xml):
            attributes = _attributes(match.group(1))
            identifier = attributes.get("Id")
            target = attributes.get("Target")
            if identifier and target:
                targets[identifier] = target
        shared: list[str] = []
        if "xl/sharedStrings.xml" in archive.namelist():
            shared_xml = archive.read("xl/sharedStrings.xml")
            shared = [
                "".join(
                    _unescape(text) for text in _TEXT_RE.findall(match.group(1))
                )
                for match in _SI_RE.finditer(shared_xml)
            ]
        sheets: dict[str, list[list[Any]]] = {}
        for match in _SHEET_RE.finditer(workbook_xml):
            attributes = _attributes(match.group(1))
            name = attributes.get("name", "")
            if name not in (NORMATIVES_SHEET, ESP_SHEET):
                continue
            target = targets.get(attributes.get("r:id", ""))
            if target is None:
                raise NormativesError(f"лист «{name}»: не найдена связь с частью пакета")
            member = _resolve_member(target)
            sheets[name] = _parse_sheet(archive.read(member), shared)
        for name in (NORMATIVES_SHEET, ESP_SHEET):
            if name not in sheets:
                raise NormativesError(f"в файле нормативов нет листа «{name}»")
        return sheets


def _parse_sheet(sheet_xml: bytes, shared: list[str]) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for row_match in _ROW_RE.finditer(sheet_xml):
        body = row_match.group(1)
        if body is None:
            rows.append([])
            continue
        cells: dict[int, Any] = {}
        for cell_match in _CELL_RE.finditer(body):
            letters = cell_match.group(1) or cell_match.group(5)
            attributes = cell_match.group(3) or cell_match.group(7) or b""
            payload = cell_match.group(4) or b""
            column = _column_index(letters.decode("ascii"))
            cell_type = b""
            type_match = re.search(rb'\st="([^"]*)"', attributes)
            if type_match:
                cell_type = type_match.group(1)
            if cell_type == b"inlineStr":
                inline = _INLINE_RE.search(payload)
                text = (
                    "".join(_unescape(part) for part in _TEXT_RE.findall(inline.group(0)))
                    if inline
                    else ""
                )
                cells[column] = text
                continue
            value_match = _VALUE_RE.search(payload)
            if value_match is None:
                cells[column] = None
                continue
            raw = _unescape(value_match.group(1))
            if cell_type == b"s":
                cells[column] = shared[int(raw)]
            elif cell_type == b"str":
                cells[column] = raw
            elif cell_type == b"b":
                cells[column] = raw == "1"
            else:
                try:
                    cells[column] = float(raw)
                except ValueError:
                    cells[column] = raw
        width = max(cells) + 1 if cells else 0
        rows.append([cells.get(index) for index in range(width)])
    return rows


def read_normative_sheets(path: str | Path) -> dict[str, list[list[Any]]]:
    workbook_path = Path(path)
    if not workbook_path.is_file():
        raise NormativesError(f"файл нормативов не найден: {workbook_path}")
    try:
        return _read_sheets_via_openpyxl(workbook_path)
    except ImportError:
        return _read_sheets_via_zip(workbook_path)


def parse_normative_set(sheets: dict[str, list[list[Any]]]) -> NormativeSet:
    values: dict[str, float] = {}
    for row in sheets[NORMATIVES_SHEET][1:]:
        if not row:
            continue
        code = row[0]
        if not isinstance(code, str):
            continue
        code = code.strip()
        field = SCALAR_CODES.get(code)
        if field is None:
            continue
        raw = row[2] if len(row) > 2 else None
        values[field] = _scale(code, _to_float(raw, code))

    missing = sorted(set(SCALAR_CODES.values()) - set(values))
    if missing:
        raise NormativesError(f"в файле нормативов нет величин: {', '.join(missing)}")

    for field, locked in METHODOLOGY_LOCKED_RUB.items():
        if values[field] != locked:
            raise NormativesError(
                f"{field}={values[field]} расходится с запертой Методикой "
                f"величиной {locked}: эталон перетирает поданное значение своим"
            )

    catalog: list[EspCatalogEntry] = []
    for row in sheets[ESP_SHEET][1:]:
        if not row or row[0] in (None, ""):
            continue
        if len(row) < 4:
            raise NormativesError(f"неполная строка таблицы ЭЦН: {row}")
        catalog.append(
            EspCatalogEntry(
                nominal=_to_float(row[0], "ЭЦН.nominal"),
                interval_low=_to_float(row[1], "ЭЦН.interval_low"),
                interval_high=_to_float(row[2], "ЭЦН.interval_high"),
                cost_rub=_to_float(row[3], "ЭЦН.cost_rub") * RUB_PER_MILLION,
            )
        )
    if not catalog:
        raise NormativesError("в файле нормативов пуста таблица ЭЦН")
    catalog.sort(key=lambda entry: entry.nominal)

    return NormativeSet(esp_catalog=tuple(catalog), **values)


def load_normatives(path: str | Path) -> NormativeSet:
    return parse_normative_set(read_normative_sheets(path))
